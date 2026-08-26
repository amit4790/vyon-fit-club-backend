"""Build the admin member Excel export."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from models import Invoice, Member, MembershipSubscription
from repositories.invoice_repository import InvoiceRepository
from repositories.member_repository import MemberRepository
from repositories.subscription_repository import SubscriptionRepository

_PLAN_DURATION_SUFFIX_RE = re.compile(
    r"\s*-\s*\d+\s*(day|days|month|months|year|years)\s*$",
    re.IGNORECASE,
)

_HEADERS = [
    "Member ID",
    "Member Name",
    "Mobile",
    "Gender",
    "Membership Name",
    "Membership Start Date",
    "Membership Cost",
    "Discount",
    "Amount Paid",
    "Outstanding",
    "Membership End Date",
    "Status",
    "Counsellor",
]


@dataclass
class MemberExportRow:
    member_id: int
    member_name: str
    mobile: str
    gender: str
    membership_name: str | None
    start_date: date | None
    membership_cost: Decimal | None
    discount: Decimal | None
    amount_paid: Decimal | None
    outstanding: Decimal | None
    end_date: date | None
    status: str
    counsellor: str | None


class MemberExportService:
    """Export members with current membership and invoice snapshot."""

    def __init__(self, db: Session):
        self.member_repo = MemberRepository(db)
        self.subscription_repo = SubscriptionRepository(db)
        self.invoice_repo = InvoiceRepository(db)

    def build_xlsx(self) -> bytes:
        rows = self._build_rows()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Members"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="B91C1C")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_index, header in enumerate(_HEADERS, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin

        money_columns = {7, 8, 9, 10}
        date_columns = {6, 11}

        for row_index, row in enumerate(rows, start=2):
            values = [
                row.member_id,
                row.member_name,
                row.mobile,
                row.gender,
                row.membership_name,
                row.start_date,
                row.membership_cost,
                row.discount,
                row.amount_paid,
                row.outstanding,
                row.end_date,
                row.status,
                row.counsellor,
            ]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=_excel_value(value))
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                if col_index in money_columns and value is not None:
                    cell.number_format = "#,##0.00"
                if col_index in date_columns and value is not None:
                    cell.number_format = "YYYY-MM-DD"

        _autosize_columns(sheet)
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{max(len(rows) + 1, 1)}"
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _build_rows(self) -> list[MemberExportRow]:
        members = self.member_repo.list_non_deleted_members()
        member_ids = [member.id for member in members]
        subscriptions = self.subscription_repo.list_subscriptions_for_member_ids(member_ids)

        subscriptions_by_member: dict[int, list[MembershipSubscription]] = {}
        for subscription in subscriptions:
            subscriptions_by_member.setdefault(subscription.member_id, []).append(subscription)

        invoices = self.invoice_repo.list_invoices_for_subscription_ids(
            [subscription.id for subscription in subscriptions]
        )
        latest_invoice_by_subscription: dict[int, Invoice] = {}
        for invoice in invoices:
            if invoice.subscription_id not in latest_invoice_by_subscription:
                latest_invoice_by_subscription[invoice.subscription_id] = invoice

        today = date.today()
        rows = [
            self._row_for_member(
                member,
                subscriptions_by_member.get(member.id, []),
                latest_invoice_by_subscription,
                today,
            )
            for member in members
        ]
        rows.sort(key=lambda row: (row.end_date is None, row.end_date or date.max, row.member_id))
        return rows

    def _row_for_member(
        self,
        member: Member,
        subscriptions: list[MembershipSubscription],
        latest_invoice_by_subscription: dict[int, Invoice],
        today: date,
    ) -> MemberExportRow:
        gender = (member.gender or "").strip()
        gender_label = gender.capitalize() if gender else ""

        if not subscriptions:
            return MemberExportRow(
                member_id=member.id,
                member_name=member.full_name,
                mobile=member.mobile_number,
                gender=gender_label,
                membership_name=None,
                start_date=None,
                membership_cost=None,
                discount=None,
                amount_paid=None,
                outstanding=None,
                end_date=None,
                status="No Membership",
                counsellor=None,
            )

        active = [
            item
            for item in subscriptions
            if item.status == "active" and item.end_date >= today
        ]
        if active:
            focus = min(active, key=lambda item: (item.end_date, item.id))
            invoice = latest_invoice_by_subscription.get(focus.id)
            money = _money_from_invoice(focus, invoice)
            return MemberExportRow(
                member_id=member.id,
                member_name=member.full_name,
                mobile=member.mobile_number,
                gender=gender_label,
                membership_name=_plan_name(focus),
                start_date=focus.start_date,
                membership_cost=money.cost,
                discount=money.discount,
                amount_paid=money.amount_paid,
                outstanding=money.outstanding,
                end_date=focus.end_date,
                status=_active_status_label(money.payment_status),
                counsellor=money.counsellor,
            )

        latest = max(subscriptions, key=lambda item: (item.end_date, item.id))
        invoice = latest_invoice_by_subscription.get(latest.id)
        money = _money_from_invoice(latest, invoice)
        return MemberExportRow(
            member_id=member.id,
            member_name=member.full_name,
            mobile=member.mobile_number,
            gender=gender_label,
            membership_name=_plan_name(latest),
            start_date=latest.start_date,
            membership_cost=money.cost,
            discount=money.discount,
            amount_paid=money.amount_paid,
            outstanding=money.outstanding,
            end_date=latest.end_date,
            status="Expired",
            counsellor=money.counsellor,
        )


@dataclass
class _MoneySnapshot:
    cost: Decimal
    discount: Decimal
    amount_paid: Decimal
    outstanding: Decimal
    payment_status: str
    counsellor: str | None


def _to_money(value: object | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _plan_name(subscription: MembershipSubscription) -> str:
    raw = ""
    if subscription.plan is not None:
        raw = subscription.plan.name or ""
    return _PLAN_DURATION_SUFFIX_RE.sub("", raw).strip()


def _money_from_invoice(subscription: MembershipSubscription, invoice: Invoice | None) -> _MoneySnapshot:
    if invoice is None:
        cost = _to_money(subscription.total_amount)
        payment_status = (subscription.payment_status or "pending").strip().lower()
        if payment_status == "paid":
            return _MoneySnapshot(
                cost=cost,
                discount=Decimal("0.00"),
                amount_paid=cost,
                outstanding=Decimal("0.00"),
                payment_status="paid",
                counsellor=None,
            )
        return _MoneySnapshot(
            cost=cost,
            discount=Decimal("0.00"),
            amount_paid=Decimal("0.00"),
            outstanding=cost,
            payment_status=payment_status or "pending",
            counsellor=None,
        )

    cost = _to_money(invoice.original_price if invoice.original_price is not None else subscription.total_amount)
    final_amount = _to_money(
        invoice.final_amount_received if invoice.final_amount_received is not None else invoice.amount
    )
    discount = (
        _to_money(invoice.discount_amount)
        if invoice.discount_amount is not None
        else max(cost - final_amount, Decimal("0.00"))
    )
    if invoice.total_paid is not None:
        amount_paid = _to_money(invoice.total_paid)
    elif invoice.amount_paid_today is not None:
        amount_paid = _to_money(invoice.amount_paid_today)
    else:
        amount_paid = Decimal("0.00")

    outstanding = (
        _to_money(invoice.outstanding_balance)
        if invoice.outstanding_balance is not None
        else max(final_amount - amount_paid, Decimal("0.00"))
    )

    invoice_status = (invoice.status or "").strip().lower()
    if outstanding <= Decimal("0.00") or invoice_status == "paid":
        payment_status = "paid"
    elif amount_paid > Decimal("0.00") or invoice_status == "partial":
        payment_status = "partial"
    else:
        payment_status = "pending"

    counsellor = (invoice.counsellor or "").strip() or None
    return _MoneySnapshot(
        cost=cost,
        discount=discount,
        amount_paid=amount_paid,
        outstanding=outstanding,
        payment_status=payment_status,
        counsellor=counsellor,
    )


def _active_status_label(payment_status: str) -> str:
    if payment_status == "paid":
        return "Active"
    if payment_status == "partial":
        return "Active - Pending Payment"
    return "Inactive (Unpaid)"


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _autosize_columns(sheet: Worksheet) -> None:
    widths = {
        1: 12,
        2: 28,
        3: 16,
        4: 12,
        5: 28,
        6: 22,
        7: 16,
        8: 12,
        9: 14,
        10: 14,
        11: 22,
        12: 26,
        13: 22,
    }
    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width
