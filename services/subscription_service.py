"""
Subscription business logic and pricing calculations.
"""

from __future__ import annotations

import io
import json
import logging
from calendar import monthrange
from dataclasses import asdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP

from sqlalchemy.orm import Session

from core.config import settings
from models import Invoice, MembershipPlan, MembershipSubscription
from repositories.invoice_repository import InvoiceRepository
from repositories.subscription_repository import SubscriptionRepository
from schemas.subscription import (
    PlanFamilyResponse,
    PlanOptionResponse,
    SubscriptionResponse,
)
from services.notification_service import NotificationService

logger = logging.getLogger(__name__)


class MemberNotFoundError(Exception):
    """Raised when a member does not exist."""


class PlanNotFoundError(Exception):
    """Raised when a membership plan does not exist."""


class SubscriptionConflictError(Exception):
    """Raised when assigning a subscription overlaps with active dates."""


class SubscriptionNotFoundError(Exception):
    """Raised when a subscription does not exist."""


@dataclass
class ExpiringResult:
    items: list[SubscriptionResponse]
    total_items: int


class SubscriptionService:
    """Business logic for membership plans and subscriptions."""

    def __init__(self, db: Session):
        self.db = db
        self.repo = SubscriptionRepository(db)
        self.invoice_repo = InvoiceRepository(db)
        self.notifier = NotificationService()

    @staticmethod
    def _to_money(value: Decimal | float | int) -> Decimal:
        return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    @staticmethod
    def _add_months(start_date: date, months: int) -> date:
        month = start_date.month - 1 + months
        year = start_date.year + month // 12
        month = month % 12 + 1
        day = min(start_date.day, monthrange(year, month)[1])
        return date(year, month, day)

    @staticmethod
    def _format_duration_label(duration_value: int, duration_unit: str) -> str:
        unit_label = "Month" if duration_unit == "months" and duration_value == 1 else (
            "Months" if duration_unit == "months" else ("Day" if duration_value == 1 else "Days")
        )
        return f"{duration_value} {unit_label}"

    @classmethod
    def _format_combined_duration_label(
        cls,
        duration_value: int,
        duration_unit: str,
        bonus_duration_value: int | None = None,
        bonus_duration_unit: str | None = None,
    ) -> str:
        paid_label = cls._format_duration_label(duration_value, duration_unit)
        if not bonus_duration_value or bonus_duration_value <= 0 or not bonus_duration_unit:
            return paid_label
        if duration_unit == bonus_duration_unit:
            if duration_unit == "months":
                unit_label = "Month" if duration_value == 1 and bonus_duration_value == 1 else "Months"
            else:
                unit_label = "Day" if duration_value == 1 and bonus_duration_value == 1 else "Days"
            return f"{duration_value} + {bonus_duration_value} {unit_label}"
        bonus_label = cls._format_duration_label(bonus_duration_value, bonus_duration_unit)
        return f"{paid_label} + {bonus_label}"

    def _calculate_end_date(self, *, start_date: date, duration_value: int, duration_unit: str) -> date:
        if duration_unit == "months":
            return self._add_months(start_date, duration_value) - timedelta(days=1)
        return start_date + timedelta(days=duration_value - 1)

    def _calculate_end_date_with_bonus(
        self,
        *,
        start_date: date,
        duration_value: int,
        duration_unit: str,
        bonus_duration_value: int | None = None,
        bonus_duration_unit: str | None = None,
    ) -> date:
        paid_end = self._calculate_end_date(
            start_date=start_date,
            duration_value=duration_value,
            duration_unit=duration_unit,
        )
        if not bonus_duration_value or bonus_duration_value <= 0 or not bonus_duration_unit:
            return paid_end
        bonus_start = paid_end + timedelta(days=1)
        return self._calculate_end_date(
            start_date=bonus_start,
            duration_value=bonus_duration_value,
            duration_unit=bonus_duration_unit,
        )

    @staticmethod
    def _normalize_bonus(
        bonus_duration_value: int | None,
        bonus_duration_unit: str | None,
    ) -> tuple[int | None, str | None]:
        if bonus_duration_value is None or bonus_duration_value <= 0 or not bonus_duration_unit:
            return None, None
        return bonus_duration_value, bonus_duration_unit

    def _resolve_duration_label(self, row: MembershipSubscription) -> str:
        if row.duration_label:
            return row.duration_label

        if row.duration_value and row.duration_unit:
            return self._format_combined_duration_label(
                row.duration_value,
                row.duration_unit,
                row.bonus_duration_value,
                row.bonus_duration_unit,
            )

        if row.plan is None:
            total_days = (row.end_date - row.start_date).days + 1
            return self._format_duration_label(max(total_days, 1), "days")

        default_end_date = self._add_months(row.start_date, row.plan.duration_months) - timedelta(days=1)
        if row.end_date == default_end_date:
            return row.plan.duration_label

        total_days = (row.end_date - row.start_date).days + 1
        return self._format_duration_label(max(total_days, 1), "days")

    def sync_expired_subscriptions(self) -> None:
        today = date.today()
        expired = self.repo.list_expired_subscriptions(today=today)
        if not expired:
            return

        affected_member_ids = {row.member_id for row in expired}
        for row in expired:
            row.status = "expired"
        self.db.commit()

        # Disable face/access on device when the member has no remaining active plan.
        self._sync_device_access_for_members(affected_member_ids)

    def _sync_device_access_for_members(self, member_ids: set[int]) -> None:
        """Queue Pri=0/1 USERINFO updates so device access matches membership."""
        if not member_ids or not settings.device_push_enabled:
            return

        from services.push_device_service import PushDeviceService

        today = date.today()
        push = PushDeviceService(self.db)
        for member_id in member_ids:
            member = self.repo.get_member_by_id(member_id)
            if not member:
                continue
            enabled = self.repo.member_has_active_membership(member_id, today)
            try:
                push.set_member_access_on_devices(
                    member_id,
                    member.full_name,
                    enabled=enabled,
                )
            except Exception:
                logger.exception(
                    "Failed to sync device access after membership change",
                    extra={"member_id": member_id, "enabled": enabled},
                )

    def get_plan_catalog(self) -> list[PlanFamilyResponse]:
        plans = self.repo.list_active_plans()

        grouped: dict[str, PlanFamilyResponse] = {}
        for plan in plans:
            family_key = plan.family_name.upper()
            if family_key not in grouped:
                includes = []
                if plan.includes_json:
                    try:
                        includes = [str(item) for item in json.loads(plan.includes_json)]
                    except json.JSONDecodeError:
                        includes = []

                grouped[family_key] = PlanFamilyResponse(
                    family=family_key,
                    description=plan.description or "",
                    includes=includes,
                    options=[],
                )

            base_price = self._to_money(plan.base_price)
            tax_percent = self._to_money(plan.tax_percent)
            tax_amount = self._to_money(base_price * tax_percent / Decimal("100"))
            total_price = self._to_money(base_price + tax_amount)

            grouped[family_key].options.append(
                PlanOptionResponse(
                    id=plan.id,
                    sku=plan.name,
                    label=plan.name,
                    variant=plan.variant_name,
                    duration_months=plan.duration_months,
                    duration_label=plan.duration_label,
                    base_price=float(base_price),
                    tax_percent=float(tax_percent),
                    tax_amount=float(tax_amount),
                    total_price=float(total_price),
                )
            )

        return list(grouped.values())

    def _plan_to_option(self, plan: MembershipPlan) -> PlanOptionResponse:
        base_price = self._to_money(plan.base_price)
        tax_percent = self._to_money(plan.tax_percent)
        tax_amount = self._to_money(base_price * tax_percent / Decimal("100"))
        total_price = self._to_money(base_price + tax_amount)

        return PlanOptionResponse(
            id=plan.id,
            sku=plan.name,
            label=plan.name,
            variant=plan.variant_name,
            duration_months=plan.duration_months,
            duration_label=plan.duration_label,
            base_price=float(base_price),
            tax_percent=float(tax_percent),
            tax_amount=float(tax_amount),
            total_price=float(total_price),
        )

    def update_plan_pricing(
        self,
        *,
        plan_id: int,
        base_price: Decimal | float,
        tax_percent: Decimal | float | None,
    ) -> PlanOptionResponse:
        plan = self.repo.get_plan_by_id(plan_id)
        if not plan:
            raise PlanNotFoundError("Membership plan not found")

        normalized_base = self._to_money(base_price)
        normalized_tax = self._to_money(tax_percent if tax_percent is not None else plan.tax_percent)
        normalized_total = self._to_money(normalized_base + (normalized_base * normalized_tax / Decimal("100")))

        plan.base_price = float(normalized_base)
        plan.tax_percent = float(normalized_tax)
        plan.total_price = float(normalized_total)
        plan.price = float(normalized_total)

        try:
            self.db.commit()
            self.db.refresh(plan)
        except Exception:
            self.db.rollback()
            raise

        return self._plan_to_option(plan)

    def assign_subscription(
        self,
        member_id: int,
        plan_id: int,
        start_date: date,
        duration_value: int | None = None,
        duration_unit: str | None = None,
        bonus_duration_value: int | None = None,
        bonus_duration_unit: str | None = None,
    ) -> tuple[SubscriptionResponse, list[dict]]:
        member = self.repo.get_member_by_id(member_id)
        if not member:
            raise MemberNotFoundError("Member not found")

        plan = self.repo.get_active_plan_by_id(plan_id)
        if not plan:
            raise PlanNotFoundError("Membership plan not found")

        # Multiple concurrent active memberships are allowed.
        base_price = self._to_money(plan.base_price)
        tax_percent = self._to_money(plan.tax_percent)
        tax_amount = self._to_money(base_price * tax_percent / Decimal("100"))
        total_amount = self._to_money(base_price + tax_amount)

        effective_duration_value = duration_value if duration_value is not None else plan.duration_months
        effective_duration_unit = duration_unit if duration_unit is not None else "months"
        effective_bonus_value, effective_bonus_unit = self._normalize_bonus(
            bonus_duration_value,
            bonus_duration_unit,
        )
        duration_label = self._format_combined_duration_label(
            effective_duration_value,
            effective_duration_unit,
            effective_bonus_value,
            effective_bonus_unit,
        )
        end_date = self._calculate_end_date_with_bonus(
            start_date=start_date,
            duration_value=effective_duration_value,
            duration_unit=effective_duration_unit,
            bonus_duration_value=effective_bonus_value,
            bonus_duration_unit=effective_bonus_unit,
        )

        row = MembershipSubscription(
            member_id=member_id,
            plan_id=plan.id,
            start_date=start_date,
            end_date=end_date,
            duration_value=effective_duration_value,
            duration_unit=effective_duration_unit,
            bonus_duration_value=effective_bonus_value,
            bonus_duration_unit=effective_bonus_unit,
            duration_label=duration_label,
            status="active",
            payment_status="pending",
            base_price=float(base_price),
            tax_percent=float(tax_percent),
            tax_amount=float(tax_amount),
            total_amount=float(total_amount),
        )

        invoice = Invoice(
            member_id=member_id,
            subscription_id=0,
            amount=float(total_amount),
            status="pending",
        )

        try:
            row = self.repo.create_subscription(row)
            invoice.subscription_id = row.id
            self.invoice_repo.create_invoice(invoice)
            self.db.commit()
            self.db.refresh(row)
            self.db.refresh(invoice)
        except Exception:
            self.db.rollback()
            raise

        # Notifications are best-effort for dummy integrations and should not fail assignment.
        welcome_results = self.notifier.send_welcome(
            member_name=member.full_name,
            email=member.email,
            phone=member.phone,
            plan_label=plan.name,
            start_date=row.start_date,
            end_date=row.end_date,
        )
        invoice_results = self.notifier.send_invoice_issued(
            member_name=member.full_name,
            email=member.email,
            phone=member.phone,
            invoice_id=invoice.id,
            amount=float(invoice.amount),
        )

        notifications = [asdict(item) for item in [*welcome_results, *invoice_results]]
        # New active plan → re-enable face/access on biometric devices.
        self._sync_device_access_for_members({member_id})
        return self._to_subscription_response(row, duration_label_override=duration_label), notifications

    def change_subscription_plan(
        self,
        subscription_id: int,
        plan_id: int,
        start_date: date | None = None,
        duration_value: int | None = None,
        duration_unit: str | None = None,
        bonus_duration_value: int | None = None,
        bonus_duration_unit: str | None = None,
    ) -> SubscriptionResponse:
        """Update an existing subscription to a different membership plan."""
        row = self.repo.get_subscription_by_id(subscription_id)
        if not row:
            raise SubscriptionNotFoundError("Subscription not found")

        if row.status not in {"active", "expired"}:
            raise SubscriptionConflictError("Only active or expired subscriptions can change plan")

        plan = self.repo.get_active_plan_by_id(plan_id)
        if not plan:
            raise PlanNotFoundError("Membership plan not found")

        effective_start = start_date or row.start_date

        effective_duration_value = duration_value if duration_value is not None else plan.duration_months
        effective_duration_unit = duration_unit if duration_unit is not None else "months"
        effective_bonus_value, effective_bonus_unit = self._normalize_bonus(
            bonus_duration_value,
            bonus_duration_unit,
        )
        duration_label = self._format_combined_duration_label(
            effective_duration_value,
            effective_duration_unit,
            effective_bonus_value,
            effective_bonus_unit,
        )
        end_date = self._calculate_end_date_with_bonus(
            start_date=effective_start,
            duration_value=effective_duration_value,
            duration_unit=effective_duration_unit,
            bonus_duration_value=effective_bonus_value,
            bonus_duration_unit=effective_bonus_unit,
        )

        base_price = self._to_money(plan.base_price)
        tax_percent = self._to_money(plan.tax_percent)
        tax_amount = self._to_money(base_price * tax_percent / Decimal("100"))
        total_amount = self._to_money(base_price + tax_amount)

        row.plan_id = plan.id
        row.start_date = effective_start
        row.end_date = end_date
        row.duration_value = effective_duration_value
        row.duration_unit = effective_duration_unit
        row.bonus_duration_value = effective_bonus_value
        row.bonus_duration_unit = effective_bonus_unit
        row.duration_label = duration_label
        row.base_price = float(base_price)
        row.tax_percent = float(tax_percent)
        row.tax_amount = float(tax_amount)
        row.total_amount = float(total_amount)
        row.status = "active"

        invoice = self.invoice_repo.get_latest_invoice_by_subscription_id(subscription_id)
        if invoice:
            total_paid = self._to_money(invoice.total_paid if invoice.total_paid is not None else 0)
            outstanding = self._to_money(max(total_amount - total_paid, Decimal("0.00")))
            invoice.amount = float(total_amount)
            invoice.original_price = float(total_amount)
            invoice.final_amount_received = float(total_amount)
            invoice.gst_amount = float(tax_amount)
            invoice.outstanding_balance = float(outstanding)
            if outstanding <= 0:
                invoice.status = "paid"
                row.payment_status = "paid"
            elif total_paid > 0:
                invoice.status = "partial"
                row.payment_status = "partial"
            else:
                invoice.status = "pending"
                row.payment_status = "pending"
                invoice.paid_at = None
        else:
            row.payment_status = "pending"

        try:
            self.db.commit()
            row = self.repo.get_subscription_by_id(subscription_id)
            if not row:
                raise SubscriptionNotFoundError("Subscription not found")
        except SubscriptionNotFoundError:
            raise
        except Exception:
            self.db.rollback()
            raise

        # Plan change / renew to active → re-enable device access when membership is valid.
        self._sync_device_access_for_members({row.member_id})
        return self._to_subscription_response(row, duration_label_override=duration_label)

    def get_member_subscriptions(self, member_id: int) -> list[SubscriptionResponse]:
        member = self.repo.get_member_by_id(member_id)
        if not member:
            raise MemberNotFoundError("Member not found")

        rows = self.repo.list_member_subscriptions(member_id)
        return [self._to_subscription_response(row) for row in rows]

    def get_expiring_subscriptions(self, days: int, page: int, page_size: int) -> ExpiringResult:
        today = date.today()
        limit_date = today + timedelta(days=days)

        rows, total_items = self.repo.list_expiring_subscriptions(
            from_date=today,
            to_date=limit_date,
            page=page,
            page_size=page_size,
        )

        return ExpiringResult(
            items=[self._to_subscription_response(row) for row in rows],
            total_items=total_items,
        )

    def build_expiring_subscriptions_xlsx(self, days: int) -> bytes:
        """Excel export of all active subscriptions ending within the lookahead window."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        today = date.today()
        limit_date = today + timedelta(days=days)
        rows, _total = self.repo.list_expiring_subscriptions(
            from_date=today,
            to_date=limit_date,
            page=1,
            page_size=10_000,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expiring"

        headers = [
            "Member ID",
            "Member Name",
            "Mobile",
            "Plan",
            "Duration",
            "Start Date",
            "End Date",
            "Days Remaining",
            "Payment Status",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="B91C1C")
        thin = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

        for row_index, row in enumerate(rows, start=2):
            member = row.member
            days_remaining = (row.end_date - today).days
            values = [
                row.member_id,
                member.full_name if member else "",
                (member.mobile_number if member else "") or "",
                row.plan.name if row.plan else "",
                self._resolve_duration_label(row),
                row.start_date.isoformat(),
                row.end_date.isoformat(),
                days_remaining,
                row.payment_status,
            ]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")

        widths = [12, 24, 16, 28, 16, 14, 14, 14, 16]
        from openpyxl.utils import get_column_letter

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _to_subscription_response(
        self,
        row: MembershipSubscription,
        duration_label_override: str | None = None,
    ) -> SubscriptionResponse:
        plan = row.plan
        duration_label = duration_label_override or self._resolve_duration_label(row)
        return SubscriptionResponse(
            id=row.id,
            member_id=row.member_id,
            member_name=row.member.full_name if row.member else None,
            plan_id=row.plan_id,
            plan_family=plan.family_name.upper(),
            plan_variant=plan.variant_name,
            plan_label=plan.name,
            duration_label=duration_label,
            duration_value=row.duration_value,
            duration_unit=row.duration_unit,
            bonus_duration_value=row.bonus_duration_value,
            bonus_duration_unit=row.bonus_duration_unit,
            start_date=row.start_date,
            end_date=row.end_date,
            status=row.status,
            base_price=float(row.base_price),
            tax_percent=float(row.tax_percent),
            tax_amount=float(row.tax_amount),
            total_amount=float(row.total_amount),
            payment_status=row.payment_status,
        )
